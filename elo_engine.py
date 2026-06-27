"""
elo_engine.py — Eggy Cup rating computation.

Reads the canonical xlsx (e.g. "Eggy Cup 87-89.xlsx"), computes Glicko-2 +
weighted ELO from the elimination-order leaderboards, writes elo_results.json
and alldata.json.

Glicko-2 percentile-based update (single observation per cup) is ported
verbatim from the ZSL project (build_zsl_glicko.py).
"""
import openpyxl, json, re, sys, os, math
from collections import defaultdict
sys.stdout.reconfigure(encoding='utf-8')

_dir = os.path.dirname(os.path.abspath(__file__))
_p = lambda f: os.path.join(_dir, f)


def parse_file(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    cups = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        position_cells = []
        for ri, row in enumerate(rows):
            for ci, val in enumerate(row):
                if val == 'Position':
                    position_cells.append((ri, ci))
        for pos_row, pos_col in position_cells:
            cup_name = None
            for sr in range(pos_row - 1, max(pos_row - 5, -1), -1):
                val = rows[sr][pos_col] if pos_col < len(rows[sr]) else None
                if val and str(val).startswith('Eggy'):
                    cup_name = str(val).strip()
                    break
            if not cup_name:
                continue
            players = []
            last_pos = None
            for row in rows[pos_row + 1:]:
                if pos_col >= len(row) or pos_col + 1 >= len(row):
                    continue
                pos, name = row[pos_col], row[pos_col + 1]
                if name is None:
                    continue
                name_str = str(name).strip()
                if name_str.startswith('*'):
                    continue
                if pos is not None:
                    try:
                        pos_clean = str(pos).rstrip('*').strip()
                        last_pos = int(float(pos_clean))
                        players.append((last_pos, name_str))
                    except Exception:
                        continue
                elif last_pos is not None:
                    players.append((last_pos, name_str))
            if players:
                cups.append({'name': cup_name, 'players': sorted(players, key=lambda x: x[0])})
    return cups


all_cups = parse_file(_p('Eggy Cup 87-94.xlsx'))


def cup_num(name):
    m = re.search(r'(\d+)', name)
    return int(m.group(1)) if m else 0


all_cups.sort(key=lambda c: cup_num(c['name']))

# Deduplicate (same cup number)
seen = set()
deduped = []
for c in all_cups:
    n = cup_num(c['name'])
    if n not in seen:
        seen.add(n)
        deduped.append(c)
all_cups = deduped

print(f"Parsed {len(all_cups)} cups")
for c in all_cups:
    print(f"  {c['name']}: {len(c['players'])} players")

# Collect all names for alias detection
all_names = set()
for cup in all_cups:
    for _, name in cup['players']:
        all_names.add(name)


def strip_tag(name):
    return re.sub(r'\[.*?\]\s*', '', name).strip()


# CANONICAL aliases — inherited wholesale from the COTD project. Players who
# never appear in Eggy still cost ~nothing to carry, and lifting the dict means
# tag-changing regulars (Wheelie, Stick, vectortrajector, MackCheesy, etc.)
# already resolve correctly. Add Eggy-specific aliases here as they appear.
CANONICAL = {
    '376': ['376.0'],
    'AndMe': ['[ORIG]AndMe16', 'AndMe14', 'AndMe15', 'AndMe16', 'AndMe17', 'Andme17', '[COMY]AndMe17', '[CSC]AndMe17'],
    'bernhard': ['[Lord] bernhard'],
    'Butter': ['[ZST] Butter'],
    'Codewalt': ['CodeWalt'],
    'DragonBoi': ['[Top3]DragonBoi'],
    'Naomi': ['[POIN]Fwogiie', '[RTR]Fwogiie', '[Tran]Fwogiie', '[frog]Fwogiie', '[RTR]Fwogiie.Kawaii', 'Fwogiie.Kawaii', 'Fwogiie', '[ASJE]Naomi', 'Naomi :3', 'Nyaomi', '[meow]Nyaomi'],
    'GuAlexItar': ['[RTR]GuAlexItar', '[CTR]GuAlexItar'],
    'Hi Im Yolo': ['[HRR]Hi Im Yolo', '[RIP]Hi Im Yolo', 'Yolo'],
    'Hydro': ['[BFP] Hydro', '[CTR]Hydro', '[RTR] Hydro', '[SLOW] Hydro', 'k2blue', '[WOW]Hydro', '[ZOMN] Hydro', '[CTR[Hydro', 'MystiCookies', 'MysticVoid', 'l3purple'],
    'I_stay_sideway': ['[RTR] I_stay_sideway', 'Istaysideway', '[RTR]Istaysideway'],
    'ITz_WillleeMan': ['ITz_Willleeman', 'Itz_WillleeMan'],
    'jandje': ['[BFP] jandje', '[CTR] jandje', '[CTR]jandje'],
    'Joking': ['[BGR] Joking'],
    'justMaki': ['[test] justMaki', '[KERN]justMaki', 'JustMaki'],
    'Kernkob': ['kernkob', '[CTR]Kernkob', 'kernbooper'],
    'L3it3R': ['L3it3r', '[CTR] L3it3R', '[CTR]L3it3R', '[CTR]L3it3r'],
    'Last': ['[dumb] Last'],
    'Lazy_Echidna': ['[NIL]Lazy_Echidna', '[TOG]Lazy_Echidna', '[TOG]Lazy_echidna', '[TOG]Lazy-Echidna', '[TOG[Lazy_Echidna', 'Lazy_Ecidna', 'Lazy Echidna'],
    'Mark': ['[RTR]Mark'],
    'Metalted': ['[ZMS] Metalted', '[ZMS]Metalted', 'Matalted'],
    'Northwind': ['Noweki'],
    'Not That Guy': ['[GANS] Not That Guy', '[RTR] Not That Guy'],
    'OccasionallyAmazingGamer': ['[CSC] OccasionallyAmazingGamer', '[CSC]OccasionallyAmazingGamer'],
    'OwlPlague': ['[CTR] OwlPlague', '[CTR]OwlPlague'],
    'Pants': ['[COLD]Pants'],
    'Phoenjx': ['[HUGS]Phoenjx'],
    'Pigbuy': ['[OREO] Pigbuy', '[OREO]Pigbuy', '[OR]Pigbuy'],
    'Principe': ['[GV] Principe'],
    'Quickracer10': ['[KURK] Quickracer10', '[AJSE] Quickracer10', 'quickracer10', '[ASJE] Quickracer10', '[CC] Quickracer10', 'quickracer', 'Quick'],
    'R0nanC': ['[CTR] R0nanC', '[CTR]R0nanC', 'R0nanc'],
    'readfreak7': ['[HRR] readfreak7', '[HRR]readfreak7', '[PFE] readfreak7'],
    'Renergy': ['[TOOB] Renergy', '[TOOB]Renergy', '[UP] Renergy', '[WSHD] Renergy', '[just] Renergy', '[meh] Renergy'],
    'rsgold': ['[OR] rsgold'],
    'rtube': ['[TOB] rtyyyyb', '[ZET] rtyyyyb', '[ZET]rtyyyyb', '[TBD] rtyyyyb', 'rtyyyyb', '[SUCK] rtube', '[TBD] rtube', '[TBD]rtube', '[BAP]rtube', '[Toob]rtube', '[dumb]rtube', '[sad]rtube', 'rtubert the farmer'],
    'Sandals': ['[CTR] Sandals'],
    'schmxrg': ['[6dog] schmxrg', '[dogg] schmxrg', '[goat] schmxrg'],
    'shadeely': ['Shadeely', '[KBR] shadeely', '[RonC] shadeely'],
    'St Nicholas': ['St Nic', 'St NIcholas', 'St. Nicholas', '[cozy]St Nicholas'],
    'TheBamboozler': ['[CTR] TheBamboozler', '[CTR]TheBamboozler'],
    'TheBestMaidens': ['[TBR]TheBestMaidens', '[WAM]TheBestMaidens', '[WAM] TheBestMaidens'],
    'TraNin': ['[P50] TraNin', '[TD38] TraNin'],
    'Tritonas1237': ['[zfwo]Tritonas1237'],
    'TwoFace': ['[CTR]TwoFace', '[CTR]Twoface'],
    'Warsnac': ['[CHR]Warcans', 'Warcans', '[BAP]Warsnac', '[CHR]Warsnac', '[old]Warsnac'],
    'zivecef': ['[WAM] zivecef', '[ZST] zivecef', '[ZST]zivecef'],
    'ZOMAN': ['[ARMS] ZOMAN', '[Bath] ZOMAN', '[Blub] ZOMAN', '[Choo] ZOMAN',
              '[DNF] Did Not DNF ZOMAN', 'DNF artist (ZOMAN)',
              '[KUNG] ZOMAN', '[Kung] ZOMAN', '[SLOW]ZOMAN', '[SNTA] ZOMAN',
              '[TOOB] ZOMAN', '[oOOo] ZOMAN'],
    'Zodiak': ['im washed [zodiakism]', 'zodiak'],
    'Cbad Cruiser': ['[CTR]Cbad Cruiser'],
    'LupensCruor': ['[CTR] LupensCruor'],
    'NathWalt': ['[CTR] NathWalt', '[CRT]NathWalt', '[CTR]NathWalt'],
    '3rdseyeview': ['[vibe]3rdseyeview'],
    'A Broken Forklift': ['[FORK] A Broken Forklift'],
    'a_random_tumbleweed': ['[ZET] a_random_tumbleweed'],
    'Axo': ['[RDX] axo'],
    'Hidef09': ['[badR] Hidef09', '[badr] Hidef09'],
    'icRS': ['[LATE] icRS'],
    'LILWOOLEY': ['[ZET]LILWOOLEY', '[ZST]LILWOOLEY'],
    'Lynhardt': ['[KBW] Lynhardt'],
    'Odist': ['[WAM]Odist'],
    'PandaMane': ['[FOV]PandaMane', '[FPV]PandaMane', '[CHEZ]PandaMane'],
    'PoopSheriff': ['[CTR] PoopSheriff', '[CTR]PoopSheriff'],
    'Psycho No. 7': ['[PTSD]Psycho No. 7'],
    'Roader': ['[BOB]Roader', '[OR] Roader'],
    'Shadynook': ['[CSC] Shadynook', '[CSC]Shadynook', '[LATE]Shadynook'],
    'Stick': ['[ZET]Stick', '[ZET] Stick'],
    'Striking Vyper': ['[CTR]Striking Vyper'],
    'XpERt': ['[TBD]XpERt'],
    'Mμ': ['[CTR]Mμ', 'Mu'],
    'Kaiser64': ['[TEA]Kaiser64'],
    'LKat': ['[GGG]Lkat', '[GGG]LKat', '[MMM]LKat'],
    'void': ['[ZET]void', '[poop]void', '[bob]void', '[popo]void', 'Void', '//////void', '[VOID[//////void', '[gorp]//////void', '[SWDN]//////void', 'timeless void', '𝒱V𝑜o𝒾i𝒹d𝒱'],
    'Achmetha0626': ['[ERR]Achmetha0626'],
    'Ax1ss': ['[FaS] Ax1ss', '[TOG] Ax1ss'],
    'BootyMcShooty88': ['[SBOI] BootyMcShooty88'],
    'Chinpokomon': ['[RB] Chinpokomon', '[RB]Chinpokomon'],
    'CopperFeather': ['[WHI] CopperFeather'],
    'DBNULL': ['[ZAGA]DBNULL'],
    'DocRee': ['[CLWN]DocRee'],
    'dudeeitsraymond': ['[iDad] dudeeitsraymond', '[iDad]dudeeitsraymond'],
    'Exterminate': ['[XTR] Exterminate', '[XTR]Exterminate'],
    'FlyBoy': ['Fly8oy', 'Flyboy', 'flyboy'],
    'Gimpel': ['Leviathan (Gimpel)'],
    'GuillaumePN': ['[Qc] GuillaumePN'],
    'incredulouspotato': ['[PINK]incredulouspotato', '[TTV]incredulouspotato'],
    'ioi8': ['[TOG][KBR] ioi8', '[TOG]ioi8'],
    'it_is_nic': ['[its] it_is_nic'],
    'Jeffrey': ['[BOGO]Jeffrey'],
    'Meowbee': ['[HRT!]Meowbee'],
    'Pheonjx': ['[HUGS]Pheonjx'],
    'Pilaf': ['[PILF] Pilaf'],
    'Sheriff': ['[Poop]Sheriff'],
    'Smullie': ['[KURK] Smullie', '[KURK]Smullie'],
    'Lilly Fenn': ['[Fenn]Lilly Fenn', 'ttv/Lilly the Bun', '[bnuy]ttv/Lilly the Bun', 'Lilly the Bun'],
    'Ulv_RaVn': ['[XTR] [VK] Ulv_RaVn', '[XTR] [VK] Ulv_Ravn'],
    'vectortrajector': ['[ZET]vectortrajector', '[ZET] vectortrajector'],
    'Victor': ['[GGG]Victor', '[MMM]Victor', '[FPV]Victor', '[RFV]Victor'],
    "Zeke Ryu'kai": ["[BoF3] Zeke Ryu'kai"],
    '=XDC=WOLF': ['=XDC=Wolf'],
    'tws20so': ['[DRFT] tws20so', '[IRS] tws20so'],
    'IronDragon111000': ['[CSC]IronDragon111000'],
    'Mokster': ['[CSC]Mokster'],
    'vortex': ['[mib]vortex'],
    'Lexer': ['[BRIT] Lexer'],
    'Murrl': ['[BAP]Murrl', '[Burp]Murrl', '[Toob]Murrl', 'MeroMeroNoMi', '[Mero] Murri'],
    'JakeAdjacent': ['[CD]SadD0ge', '[SWMG]SadD0ge', '[SWMG]JakeAdjacent', 'SadD0ge'],
    'Beans': ['[CTR]Beans'],
    'DeeDeeNaNaNa': ['[CSC] DeeDeeNaNaNa', '[CSC]DeeDeeNaNaNa'],
    'Form': ['[fn]Form'],
    'Hellmet': ['[Dark]Hellmet', '[ZOFC]Hellmet'],
    'Shattersmith': ['Jakie', '[CD] Jakie', '[ZET] Jakie', 'shattersmith'],
    'K410K3N': ['[20X]K410K3N', '[20x]K410K3N', '[Gwen]K410K3N'],
    'loganbradley714': ['[GFHL]loganbradley714', '[GLHF]loganbradley714'],
    'lucanakin': ['[DNFF]lucanakin'],
    'MarcSubstitute': ['[DHLU]MarcSubstitute', '[SLOW]MarcSubstitute'],
    'MetalCJ': ['[TTR]MetalCJ'],
    'microways': ['[Quac] microways'],
    'MMXD18': ['[Toob]MMXD18'],
    'Moody': ['[CTR]Moody', '[MIB]Moody'],
    'RadAbsRad': ['[Meow]RadAbsRad'],
    'Redstony': ['[Stc3]Redstony', '[TILT]Redstony'],
    'Six': ['SixSixSevenSeven', '[BAP]SixSixSevenSeven', '[BAP]Six'],
    'Sterben': ['[BAP]Sterben', '[PNCK]Mini P.E.K.K.A', 'λ', 'Lλmbda', '[FPV]Lλmbda', '[PCDJ]Sterben'],
    'stindt': ['[KAAS]stindt', '[Lame]stindt', '[Lame}stindt', '[Same]stindt', '[Slow]stindt', '[Tame]stindt'],
    'Zeus': ['[NewB]Zeus', '[NewB] Zeus', '[SLOW]Zeus'],
    'Tommygaming': ['[CSC]Tommygaming', '[OOPS]Tommygaming', '[jofk]Tommygaming', 'TommyGaming5132', 'Tommygaming5132', '[TG]Tommygaming5132', '[CSC]Tommygaming5132', '[CSC]Tommygaming6132', '[C3PO]Tommygaming5132'],
    'WotterBytes': ['Wotterbytes'],
    'ping': ['[bad] ping', '[boom] ping', '[no] ping', '[pong]ping'],
    'agix': ['[GYMC] agix', '[CTR]agix', '[CTR] agix'],
    'An Actual g00se': ['g00se', '[CSC] An Actual g00se', '[Err] An Actual g00se', '[CSC] BaBa is g00se', '[CSC] CantFindTheg00se'],
    'BOB THE GAMER': ['[BOGO]BOB THE GAMER', '[MEAT]BOB THE GAMER'],
    'ferinine': ['[Err]ferinine', '[Err]ferninine', '[ERR]ferinine'],
    'frenchteost': ['[LEXR]frenchteost', '[ZET]frenchteost'],
    'Ionjig': ['ionjig'],
    'Lamp': ['[CTR]Lamp', '[The]Lamp', '[bam]Lamp'],
    'MackCheesy': ['[CHEZ]MackCheesy', '[ZET]MackCheesy'],
    'magostinho20': ['[F1] magostinho20', '[I290] magostinho20'],
    'MeOne2Three4': ['meone2three4'],
    'Mortishade': ['[CTR]Mortishade', '[bam]Mortishade'],
    'Mr. Hubub': ['[Heyo]Mr. Hubub', '[heyo]Mr. Hubub'],
    'OLR94': ['[CSC]OLR94'],
    'Ploddip': ['[ZET] Ploddip', '[ZET]Ploddip'],
    'redal': ['[CSC] redal'],
    'SkyVirus': ['[NOOB]SkyVirus'],
    'Socks242': ['[Fly] Socks242'],
    'variableferret': ['[CSC] variableferret'],
    'Weak_Knees': ['[COMY]Weak_Knees', 'Weak_knees', '[Burp]Weak_Knees'],
    'Wheelie': ['[ZET] Wheelie', '[ZET]Wheelie'],
    'Zachafinackus': ['[Sumo]Zachafinackus'],
    'Heart-TGV': ['[TTR]Heart-TGV'],
    'sailingman': ['segelnhoch3'],
    'captancraft2': ['[PINK] captancraft2'],
    # ── From Kerki canonical (Kerki/Cross-Comp regulars who also show up in Eggy) ──
    'aizpun': ['[KURK]aizpun', '[KURK] aizpun'],
    'brrryy': ['brryyy', 'brrrryy'],
    'DorthJohson': ['Dorth Johson', 'DorthJohnson'],
    'Eclipse135': ['Eclipse125'],
    'JobW': ['Job'],
    'Jinx': ['[DCS] Jinx'],
    'LArk': ['[MMM]LArk', 'Lark', '[MMM]Lark'],
    'lil_zwimpie': ['[KURK]lil_zwimpie', '[KURK] lil_zwimpie'],
    'logix': ['[CT]logix', '[CTR]logix'],
    'LoudSentinel': ['[EUB]LoudSentinel'],
    'MrBunny_666': ['MrBunny666', 'Mr Bunny 666'],
    'PlusMicron': ['plusmicron'],
    'SharKy': ['Sharky'],
    'Tudge Boat': ['Tudge'],
    'Unfortunate Inc': ['UnfortunateInc'],
    # ── Eggy-only seeds (from Eggy logs) ──
    'maskeddog': ['[ZET]maskeddog', '[ZET] maskeddog'],
    'Shinikage221': ['[CCC]Shinikage221', '[CCC] Shinikage221'],
    'bjenk4': ['[AUz] bjenk4', '[AUz]bjenk4'],
}

NAME_MAP = {}
for canonical, aliases in CANONICAL.items():
    for alias in aliases:
        NAME_MAP[alias] = canonical

# Auto-extend: [TAG]Name where Name already exists as canonical
for n in all_names:
    stripped = strip_tag(n)
    if stripped != n and stripped in all_names and n not in NAME_MAP:
        if stripped in CANONICAL:
            NAME_MAP[n] = stripped


def normalize(name):
    return NAME_MAP.get(name, name)


for cup in all_cups:
    cup['players'] = [(pos, normalize(name)) for pos, name in cup['players']]

# Warn about remaining duplicates
by_stripped = defaultdict(set)
for cup in all_cups:
    for _, name in cup['players']:
        by_stripped[strip_tag(name).lower()].add(name)
remaining = {k: v for k, v in by_stripped.items() if len(v) > 1}
if remaining:
    print(f"\nWARNING: {len(remaining)} unresolved duplicates:")
    for k, v in sorted(remaining.items()):
        print(f"  {k}: {sorted(v)}")
else:
    print("\nAll duplicates resolved.")

unique = set()
for cup in all_cups:
    for _, name in cup['players']:
        unique.add(name)
print(f"Unique players: {len(unique)}")

# === ELO ===
STARTING = 1500
K_BASE = 32
PROV_CUPS = 12
PROV_MULT = 1.5
DECAY = 0.995
GRACE = 3


def E(ra, rb):
    return 1.0 / (1.0 + 10.0 ** ((rb - ra) / 400.0))


def pct_mult(pos, n):
    pct = pos / n
    if pct <= 0.08: return 3.0
    if pct <= 0.15: return 2.0
    if pct <= 0.25: return 1.3
    if pct <= 0.50: return 0.8
    return 0.5


# ── Glicko-2 constants (ported from ZSL build_zsl_glicko.py) ──
G2_R0 = 1500.0
G2_RD0 = 350.0
G2_RD_FLOOR = 80.0
G2_VOL0 = 0.06
G2_TAU = 0.8
G2_EPSILON = 1e-6
G2_SCALE = 173.7178
G2_MAX_ITER = 100


def _g2_to_mu(r):  return (r - 1500) / G2_SCALE
def _g2_to_phi(rd): return rd / G2_SCALE
def _g2_from_mu(mu):  return mu * G2_SCALE + 1500
def _g2_from_phi(phi): return phi * G2_SCALE
def _g2_g(phi):       return 1.0 / math.sqrt(1 + 3 * phi ** 2 / (math.pi ** 2))


def _g2_E(mu, mu_j, phi_j):
    x = -_g2_g(phi_j) * (mu - mu_j)
    if x > 700:  return 0.0
    if x < -700: return 1.0
    return 1.0 / (1 + math.exp(x))


def _g2_new_vol(sigma, phi, delta, v):
    a = math.log(sigma ** 2)
    tau2 = G2_TAU ** 2

    def f(x):
        ex = math.exp(x)
        p2v = phi ** 2 + v + ex
        return (ex * (delta ** 2 - phi ** 2 - v - ex)) / (2 * p2v ** 2) - (x - a) / tau2

    A = a
    if delta ** 2 > phi ** 2 + v:
        B = math.log(delta ** 2 - phi ** 2 - v)
    else:
        k = 1
        while f(a - k * G2_TAU) < 0 and k < G2_MAX_ITER:
            k += 1
        B = a - k * G2_TAU
    fA, fB = f(A), f(B)
    for _ in range(G2_MAX_ITER):
        if abs(B - A) < G2_EPSILON:
            break
        C = A + (A - B) * fA / (fB - fA)
        fC = f(C)
        if fC * fB <= 0:
            A, fA = B, fB
        else:
            fA /= 2
        B, fB = C, fC
    return math.exp(A / 2)


def compute_glicko2(cups):
    state = {}   # name → (rating, RD, vol)
    gp = defaultdict(int)
    history = defaultdict(list)
    wins = defaultdict(int)
    pods = defaultdict(lambda: [0, 0, 0])
    best = defaultdict(lambda: 999)
    total_pos = defaultdict(int)
    avg_cups_ = defaultdict(int)
    peak = defaultdict(lambda: -9999)

    for cup in cups:
        players = cup['players']
        n = len(players)
        if n < 2:
            continue

        # Inactivity bump: phi_star = sqrt(phi^2 + vol^2). Floor at RD_FLOOR.
        for nm in state:
            r, rd, vol = state[nm]
            phi = _g2_to_phi(rd)
            phi_star = math.sqrt(phi ** 2 + vol ** 2)
            new_rd = max(min(_g2_from_phi(phi_star), G2_RD0), G2_RD_FLOOR)
            state[nm] = (r, new_rd, vol)

        # Initialize newcomers
        for _, nm in players:
            if nm not in state:
                state[nm] = (G2_R0, G2_RD0, G2_VOL0)

        # Percentile-based update (one observation per player per cup).
        new_states = {}
        for i in range(n):
            pi, ni = players[i]
            r_i, rd_i, vol_i = state[ni]
            mu_i = _g2_to_mu(r_i)
            phi_i = _g2_to_phi(rd_i)
            actual = (n - pi) / (n - 1)
            e_sum = 0.0
            g_sum = 0.0
            for j in range(n):
                if i == j: continue
                _, nj = players[j]
                r_j, rd_j, _ = state[nj]
                e_sum += _g2_E(mu_i, _g2_to_mu(r_j), _g2_to_phi(rd_j))
                g_sum += _g2_g(_g2_to_phi(rd_j))
            e_avg = e_sum / (n - 1)
            g_eff = g_sum / (n - 1)
            v_inv = g_eff ** 2 * e_avg * (1 - e_avg)
            v = 1.0 / v_inv if v_inv > 1e-10 else 1e6
            delta_sum = g_eff * (actual - e_avg)
            delta = v * delta_sum
            new_vol = _g2_new_vol(vol_i, phi_i, delta, v)
            phi_star = math.sqrt(phi_i ** 2 + new_vol ** 2)
            new_phi = 1.0 / math.sqrt(1.0 / phi_star ** 2 + 1.0 / v)
            new_mu = mu_i + new_phi ** 2 * delta_sum
            new_r = _g2_from_mu(new_mu)
            new_rd = max(_g2_from_phi(new_phi), G2_RD_FLOOR)
            new_states[ni] = (new_r, new_rd, new_vol)

        for pos, name in players:
            state[name] = new_states[name]
            r = state[name][0]
            rd = state[name][1]
            gp[name] += 1
            history[name].append({'cup': cup['name'], 'position': pos,
                                  'rating': round(r, 1), 'rd': round(rd, 1),
                                  'lobby_size': n})
            total_pos[name] += pos
            avg_cups_[name] += 1
            if pos < best[name]:
                best[name] = pos
            if r > peak[name]: peak[name] = r
            if pos == 1:   wins[name] += 1; pods[name][0] += 1
            elif pos == 2: pods[name][1] += 1
            elif pos == 3: pods[name][2] += 1

    return {'state': state, 'gp': gp, 'history': history, 'wins': wins,
            'pods': pods, 'best': best, 'total_pos': total_pos,
            'avg_cups': avg_cups_, 'peak': peak}


def compute_weighted_elo(cups):
    w_ratings = defaultdict(lambda: STARTING)
    w_gp = defaultdict(int)
    w_history = defaultdict(list)
    for cup in cups:
        players = cup['players']
        n = len(players)
        if n < 2:
            continue
        avg_field = sum(w_ratings[nm] for _, nm in players) / n
        w_cup_deltas = defaultdict(float)
        for i in range(n):
            pi, ni = players[i]
            ra = w_ratings[ni]
            for j in range(n):
                if i == j: continue
                pj, nj = players[j]
                e = E(ra, w_ratings[nj])
                s = 1.0 if pi < pj else (0.0 if pi > pj else 0.5)
                win_pos = pi if pi <= pj else pj
                win_name = ni if pi <= pj else nj
                pair_quality = (w_ratings[ni] + w_ratings[nj]) / (2 * avg_field)
                k = K_BASE / (n - 1) * pct_mult(win_pos, n) * pair_quality
                if w_gp[win_name] < PROV_CUPS:
                    k *= PROV_MULT
                w_cup_deltas[ni] += k * (s - e)
        for pos, name in players:
            w_ratings[name] += w_cup_deltas[name]
            w_gp[name] += 1
            w_history[name].append({'cup': cup['name'], 'position': pos, 'rating': round(w_ratings[name], 1), 'lobby_size': n})
    return {'ratings': w_ratings, 'gp': w_gp, 'history': w_history}


def build_weighted_list(elo_data, stat_data, cups_list, min_cups=1):
    """Weighted ELO leaderboard with inactivity decay (active = decayed raw)."""
    rat = elo_data['ratings']
    hist = elo_data['history']
    gp_d = stat_data['gp']
    best_d = stat_data['best']
    total_pos_d = stat_data['total_pos']
    avg_cups_d = stat_data['avg_cups']
    wins_d = stat_data['wins']
    pods_d = stat_data['pods']
    total_n = len(cups_list)
    last_idx = {}
    for idx, cup in enumerate(cups_list):
        for _, name in cup['players']:
            last_idx[name] = idx

    def dec(rating, name):
        missed = total_n - 1 - last_idx.get(name, 0)
        if missed <= GRACE:
            return round(rating, 1)
        return round(1500 + (rating - 1500) * (DECAY ** (missed - GRACE)), 1)

    out = []
    for name in rat:
        has_pod = sum(pods_d[name]) > 0
        if gp_d[name] < min_cups and not has_pod:
            continue
        raw = round(rat[name], 1)
        act = dec(rat[name], name)
        avg = round(total_pos_d[name] / avg_cups_d[name], 1) if avg_cups_d[name] > 0 else 0
        peak = max(h['rating'] for h in hist[name]) if hist[name] else raw
        h_list = [{'c': cup_num(h['cup']), 'r': h['rating'], 'p': h['position']} for h in hist[name]]
        out.append({
            'n': name, 'a': act, 'r': raw,
            'c': gp_d[name], 'b': best_d[name] if best_d[name] < 999 else 0,
            'v': avg, 'w': wins_d[name],
            'g': pods_d[name][0], 's': pods_d[name][1], 'z': pods_d[name][2],
            'p': round(peak, 1), 'h': h_list
        })
    out.sort(key=lambda p: p['a'], reverse=True)
    return out


def build_glicko_list(g_data, min_cups=1):
    """Glicko-2 leaderboard. No ELO-style decay — RD already captures uncertainty.
    Sort by rating (not LB95) so a tiny pool with high RD still produces a useful order."""
    state = g_data['state']
    hist = g_data['history']
    gp_d = g_data['gp']
    best_d = g_data['best']
    total_pos_d = g_data['total_pos']
    avg_cups_d = g_data['avg_cups']
    wins_d = g_data['wins']
    pods_d = g_data['pods']
    peak_d = g_data['peak']

    out = []
    for name, (rating, rd, vol) in state.items():
        has_pod = sum(pods_d[name]) > 0
        if gp_d[name] < min_cups and not has_pod:
            continue
        raw = round(rating, 1)
        avg = round(total_pos_d[name] / avg_cups_d[name], 1) if avg_cups_d[name] > 0 else 0
        peak = peak_d[name] if peak_d[name] != -9999 else raw
        h_list = [{'c': cup_num(h['cup']), 'r': h['rating'], 'rd': h['rd'], 'p': h['position']} for h in hist[name]]
        out.append({
            'n': name, 'a': raw, 'r': raw,
            'rd': round(rd, 1), 'vol': round(vol, 4),
            'lb95': round(rating - 2 * rd, 1),
            'c': gp_d[name], 'b': best_d[name] if best_d[name] < 999 else 0,
            'v': avg, 'w': wins_d[name],
            'g': pods_d[name][0], 's': pods_d[name][1], 'z': pods_d[name][2],
            'p': round(peak, 1), 'h': h_list
        })
    out.sort(key=lambda p: p['a'], reverse=True)
    return out


# --- Compute both variants ---
print("\nComputing Glicko-2...")
g_full = compute_glicko2(all_cups)
print("Computing weighted ELO...")
w_full = compute_weighted_elo(all_cups)

# --- Console output (weighted ratings, Glicko stats) ---
ratings = w_full['ratings']
gp = g_full['gp']
history = w_full['history']
wins = g_full['wins']
pods = g_full['pods']
best = g_full['best']
total_pos = g_full['total_pos']
avg_cups_s = g_full['avg_cups']

lb = sorted([(n, round(ratings[n], 1), gp[n], wins[n], pods[n], best[n], total_pos[n], avg_cups_s[n]) for n in ratings],
            key=lambda x: x[1], reverse=True)

print("\n" + "=" * 105)
print(f"{'#':<5}{'Player':<26}{'Elo':<9}{'Cups':<6}{'W':<4}{'Pod':<11}{'Avg':<7}{'Peak':<9}{'Best'}")
print("=" * 105)
for rank, (name, rating, cp, w, pd, bf, tp, ac) in enumerate(lb, 1):
    peak = max(h['rating'] for h in history[name]) if history[name] else rating
    avg = tp / ac if ac > 0 else 0
    print(f"{rank:<5}{name:<26}{rating:<9}{cp:<6}{w:<4}{pd[0]}/{pd[1]}/{pd[2]:<7}{avg:<7.1f}{peak:<9}{bf}")
    if rank >= 40: break

print(f"\nTotal: {len(lb)} | 2+: {sum(1 for _, _, g, _, _, _, _, _ in lb if g >= 2)} | 3+: {sum(1 for _, _, g, _, _, _, _, _ in lb if g >= 3)}")

# elo_results.json (used by build_cups.py to invert into cups.json)
output = {
    'parameters': {'starting_rating': STARTING, 'k_base': K_BASE, 'provisional_cups': PROV_CUPS,
                   'provisional_multiplier': PROV_MULT, 'cups_processed': len(all_cups)},
    'leaderboard': [
        {'rank': i + 1, 'name': name, 'rating': rating, 'cups': cp, 'wins': w,
         'podiums': {'gold': pd[0], 'silver': pd[1], 'bronze': pd[2]},
         'avg_position': round(tp / ac, 1) if ac > 0 else 0, 'best_finish': bf,
         'peak_rating': max(h['rating'] for h in history[name]) if history[name] else rating,
         'history': history[name]}
        for i, (name, rating, cp, w, pd, bf, tp, ac) in enumerate(lb)
    ]
}
with open(_p('elo_results.json'), 'w') as f:
    json.dump(output, f, indent=2)
print("elo_results.json saved")

# alldata.json (frontend payload — glicko + weighted)
alldata = {
    'glicko':   build_glicko_list(g_full,                       min_cups=1),
    'weighted': build_weighted_list(w_full, g_full, all_cups,   min_cups=1),
}
with open(_p('alldata.json'), 'w') as f:
    json.dump(alldata, f, separators=(',', ':'))
print(f"alldata.json written (glicko: {len(alldata['glicko'])}, weighted: {len(alldata['weighted'])} players)")
