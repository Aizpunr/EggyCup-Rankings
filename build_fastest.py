"""
build_fastest.py — extract fastest times from Eggy xlsx(es) into fastest.json.
"""
import json, os, re, sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

base = os.path.dirname(os.path.abspath(__file__))
def _p(f): return os.path.join(base, f)


def load_aliases():
    """Parse CANONICAL dict from elo_engine.py source."""
    name_map = {}
    lines = open(_p('elo_engine.py'), encoding='utf-8').readlines()
    collecting = False
    buf = []
    for line in lines:
        if not collecting and re.match(r'^CANONICAL\s*=\s*\{', line):
            collecting = True
        if collecting:
            buf.append(line)
            if line.strip() == '}':
                break
    if not buf:
        return name_map
    block = ''.join(buf).split('=', 1)[1].strip()
    canonical = eval(block)
    for canon, aliases in canonical.items():
        for alias in aliases:
            name_map[alias] = canon
    return name_map


NAME_MAP = load_aliases()


def normalize_name(name):
    if name in NAME_MAP:
        return NAME_MAP[name]
    stripped = re.sub(r'^\[.*?\]\s*', '', name).strip()
    if stripped in NAME_MAP:
        return NAME_MAP[stripped]
    return stripped


# Auto-detect xlsx files from elo_engine.py
_elo_src = open(_p('elo_engine.py'), encoding='utf-8').read()
FILES = re.findall(r"parse_file\(_p\('(.+?\.xlsx)'\)\)", _elo_src)

RE_FT = re.compile(
    r'Fastest Time:\s*'
    r'(?:(\d+):)?'
    r'([\d.]+)'
    r'\s+by\s+'
    r'(.+?)'
    r'(?:\s+in\s+(.+))?$'
)


def parse_time(minutes, seconds):
    t = float(seconds)
    if minutes:
        t += int(minutes) * 60
    return round(t, 3)


def find_cup_id(ws, row, col):
    for cc in range(col, max(0, col - 6), -1):
        for rr in (2, 3, 1):
            val = ws.cell(rr, cc).value
            if val and isinstance(val, str):
                val = val.strip()
                if 'Eggy' in val:
                    return val
    return None


def find_max_round(ws, ft_row, ft_col):
    pos_col = None
    header_row = None
    for rr in range(ft_row, ft_row + 3):
        for cc in range(max(1, ft_col - 4), ft_col + 2):
            v = ws.cell(rr, cc).value
            if v and str(v).strip() == 'Position':
                pos_col = cc
                header_row = rr
                break
        if pos_col:
            break
    if not pos_col:
        return None
    eround_col = pos_col + 3
    max_r = 0
    for rr in range(header_row + 1, ws.max_row + 1):
        name = ws.cell(rr, pos_col + 1).value
        if name is None:
            break
        er = ws.cell(rr, eround_col).value
        if er and isinstance(er, (int, float)):
            max_r = max(max_r, int(er))
    return max_r if max_r > 0 else None


def scan_file(filepath):
    results = []
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        print(f"  SKIP {os.path.basename(filepath)}: {e}")
        return results
    for sname in wb.sheetnames:
        ws = wb[sname]
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                val = ws.cell(r, c).value
                if not val or not isinstance(val, str) or 'Fastest Time:' not in val:
                    continue
                m = RE_FT.match(val)
                if not m:
                    continue
                minutes, seconds, player, round_info = m.groups()
                time_val = parse_time(minutes, seconds)
                player = player.strip()
                round_info = round_info.strip() if round_info else None
                cup_id = find_cup_id(ws, r, c) or sname
                max_round = find_max_round(ws, r, c)
                players_left = None
                if round_info == 'Final':
                    players_left = 2
                elif max_round and round_info:
                    rm = re.search(r'(\d+)', round_info)
                    if rm:
                        ft_round = int(rm.group(1))
                        players_left = max_round - ft_round + 2
                entry = {
                    'cup': cup_id,
                    'time': time_val,
                    'player': normalize_name(player),
                    'round': round_info,
                }
                if players_left:
                    entry['left'] = players_left
                results.append(entry)
    wb.close()
    return results


all_entries = []
for fname in FILES:
    path = _p(fname)
    if not os.path.exists(path):
        print(f"  WARNING: {fname} not found, skipping")
        continue
    entries = scan_file(path)
    print(f"  {fname}: {len(entries)} fastest times")
    all_entries.extend(entries)

print(f"\nTotal raw entries: {len(all_entries)}")

by_cup = {}
for e in all_entries:
    cup = e['cup']
    if cup not in by_cup:
        by_cup[cup] = e
    else:
        existing = by_cup[cup]
        has_more = (e.get('round') and not existing.get('round')) or \
                   (e.get('left') and not existing.get('left'))
        if has_more:
            by_cup[cup] = e
        elif e.get('round') and existing.get('round') and e['time'] < existing['time']:
            by_cup[cup] = e

def _cup_num(e):
    m = re.search(r'(\d+)', e['cup'])
    return int(m.group(1)) if m else -1


# Fastest lap is map-dependent, so don't rank across cups — list per event, newest first.
fastest = sorted(by_cup.values(), key=_cup_num, reverse=True)
print(f"After dedup: {len(fastest)} cups with fastest times")

with open(_p('fastest.json'), 'w', encoding='utf-8') as f:
    json.dump(fastest, f, separators=(',', ':'), ensure_ascii=False)

print(f"Wrote fastest.json ({len(fastest)} entries)")
print("\nFastest lap per cup (newest first):")
for e in fastest:
    rd = e['round'] or '?'
    print(f"  {e['cup']:<10} {e['time']:.3f}s  {e['player']:<25} {rd}")
