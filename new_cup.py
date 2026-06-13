"""
new_cup.py — Process an Eggy Cup from COTDTracker mod logs.

Usage:
  python new_cup.py 89                       # mapper unknown / TBD
  python new_cup.py 89 "MapperName"          # mapper known
  python new_cup.py 89 "MapperName" --exclude wheelie,other
  python new_cup.py 89 --from-log eggy_89.log   # parse an archived log (default)

By default reads `cup logs/eggy_<N>.log` — already-archived. If you want to
process a fresh log straight from BepInEx, pass `--live`.

Parses log → appends to xlsx → JSON backup → runs elo_engine + build_cups +
build_fastest + analyze_cup_livelog (when livelog is present).
"""
import re, os, sys, json, subprocess, shutil
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')  # players use exotic unicode names

_dir = os.path.dirname(os.path.abspath(__file__))
_p = lambda f: os.path.join(_dir, f)

LIVE_LOG_PATH = r"C:\Program Files (x86)\Steam\steamapps\common\Zeepkist\BepInEx\LogOutput.log"
LIVE_LIVELOG_PATH = r"C:\Program Files (x86)\Steam\steamapps\common\Zeepkist\BepInEx\LiveLeaderboardLogger.log"
COLS_PER_CUP = 6

# ── Parse arguments ──
if len(sys.argv) < 2:
    print("Usage: python new_cup.py <cup_number> [mapper] [--exclude name1,name2] [--live]")
    sys.exit(1)

cup_num = int(sys.argv[1])
mapper = "TBD"
for arg in sys.argv[2:]:
    if not arg.startswith('--') and mapper == "TBD":
        mapper = arg
        break

cup_id = f"Eggy {cup_num}"

extra_excluded = []
if '--exclude' in sys.argv:
    idx = sys.argv.index('--exclude')
    if idx + 1 < len(sys.argv):
        extra_excluded = [n.strip() for n in sys.argv[idx + 1].split(',') if n.strip()]

use_live = '--live' in sys.argv

excluded = ({mapper} if mapper and mapper != "TBD" else set()) | set(extra_excluded)

print(f"Processing {cup_id}, mapper: {mapper}")
if extra_excluded:
    print(f"Also excluding: {', '.join(extra_excluded)}")
print()

# ── 1. Locate / archive log ──
log_dir = _p('cup logs')
os.makedirs(log_dir, exist_ok=True)
log_backup = os.path.join(log_dir, f'eggy_{cup_num}.log')
live_log_backup = os.path.join(log_dir, f'eggy_{cup_num}_liveleaderboard.log')

if use_live:
    shutil.copy2(LIVE_LOG_PATH, log_backup)
    print(f"Live BepInEx log archived to {log_backup}")
    if os.path.exists(LIVE_LIVELOG_PATH):
        shutil.copy2(LIVE_LIVELOG_PATH, live_log_backup)
        print(f"Live livelog archived to {live_log_backup}")
    else:
        live_log_backup = None
else:
    if not os.path.exists(log_backup):
        print(f"ERROR: {log_backup} not found. Pass --live to grab from BepInEx, or save_log first.")
        sys.exit(1)
    print(f"Reading archived log: {log_backup}")
    if not os.path.exists(live_log_backup):
        live_log_backup = None

with open(log_backup, encoding='utf-8', errors='replace') as f:
    lines = [l for l in f.readlines() if 'COTDTracker' in l]

if not lines:
    print("ERROR: No COTDTracker lines found in log file.")
    sys.exit(1)

# ── 2. Parse rounds ──
rounds = []
current_round = []
for line in lines:
    if 'Doing eliminations with leaderboard' in line:
        if current_round:
            rounds.append(current_round)
        current_round = []
    elif 'Eliminating ' in line or 'Player ' in line:
        current_round.append(line)
if current_round:
    rounds.append(current_round)

if not rounds:
    print("ERROR: No elimination rounds found in log.")
    sys.exit(1)

elim_order = []
actual_round = 0
winner_marker = None  # name extracted from a final single-player "elimination" round
for rnd in rounds:
    player_times = {}
    eliminated_names = []
    for line in rnd:
        m = re.search(r'Player (.+?): Time: (.+)', line)
        if m:
            name = m.group(1).strip()
            time_str = m.group(2).strip()
            player_times[name] = time_str
        m2 = re.search(r'Eliminating (?:DNF|on time): (.+)', line)
        if m2:
            name = m2.group(1).strip()
            if name not in eliminated_names:
                eliminated_names.append(name)
    if not eliminated_names:
        continue
    # Eggy cup-end marker: only one player in leaderboard, that same player
    # "eliminated" by the mod. They're the winner — don't count as elimination.
    if len(player_times) == 1 and len(eliminated_names) == 1 \
            and next(iter(player_times.keys())) == eliminated_names[0]:
        winner_marker = eliminated_names[0]
        continue
    actual_round += 1
    for name in eliminated_names:
        if name not in excluded:
            elim_round_time = player_times.get(name, 'DNF')
            dnf = (elim_round_time == 'DNF')
            elim_order.append((name, elim_round_time, actual_round, dnf))

# Winner — prefer the explicit marker if present, otherwise survivor (COTD style)
all_named = set()
for rnd in rounds:
    for line in rnd:
        m = re.search(r'Player (.+?): Time:', line)
        if m:
            all_named.add(m.group(1).strip())
elim_set = {e[0] for e in elim_order}
if winner_marker and winner_marker not in excluded:
    winner = winner_marker
else:
    winners = [n for n in all_named if n not in elim_set and n not in excluded]
    if not winners:
        print("ERROR: Could not determine winner.")
        sys.exit(1)
    winner = winners[0]

winner_time = None
for line in reversed(lines):
    m = re.search(r'Player ' + re.escape(winner) + r': Time: (.+)', line)
    if m:
        winner_time = m.group(1).strip()
        break

# Build leaderboard: per-round finishers ordered by time, DNFs tie at bottom of round
elim_order.reverse()
leaderboard = [(winner, winner_time, None, 1)]
pos = 2
i = 0
while i < len(elim_order):
    rnd = elim_order[i][2]
    group = []
    while i < len(elim_order) and elim_order[i][2] == rnd:
        group.append(elim_order[i])
        i += 1
    finishers = []
    dnfs = []
    for name, display_time, r, dnf in group:
        if dnf:
            dnfs.append((name, display_time, r))
        else:
            try:
                t = float(str(display_time).replace(',', '.'))
                finishers.append((name, display_time, r, t))
            except (ValueError, TypeError):
                dnfs.append((name, display_time, r))
    finishers.sort(key=lambda x: x[3])
    for name, display_time, r, _ in finishers:
        leaderboard.append((name, display_time, r, pos))
        pos += 1
    if dnfs:
        dnf_pos = pos
        for name, display_time, r in dnfs:
            leaderboard.append((name, display_time, r, dnf_pos))
        pos += len(dnfs)

print(f"Parsed {len(leaderboard)} players (excluded: {', '.join(sorted(excluded)) or 'none'})")
print(f"Winner: {winner} ({winner_time})")
print(f"Rounds in log: {len(rounds)}")
print()

# Fastest time
fastest_time = None
fastest_name = None
fastest_round = None
actual_round = 0
for rnd in rounds:
    has_elim = any(re.search(r'Eliminating (?:DNF|on time):', line) for line in rnd)
    if has_elim:
        actual_round += 1
    rnd_num = actual_round if has_elim else 0
    for line in rnd:
        m = re.search(r'Player (.+?): Time: (.+)', line)
        if m:
            name = m.group(1).strip()
            time_str = m.group(2).strip()
            if time_str != 'DNF':
                t = float(time_str.replace(',', '.'))
                if fastest_time is None or t < fastest_time:
                    fastest_time = t
                    fastest_name = name
                    fastest_round = rnd_num

# ── 3. Auto-detect xlsx + insert column ──
elo_py_path = _p('elo_engine.py')
with open(elo_py_path, encoding='utf-8') as f:
    elo_src = f.read()

xlsx_matches = re.findall(r"parse_file\(_p\('(Eggy Cup \d+-\d+\.xlsx)'\)\)", elo_src)
if not xlsx_matches:
    print("ERROR: Could not find Eggy Cup xlsx reference in elo_engine.py")
    sys.exit(1)
current_xlsx = xlsx_matches[-1]
xlsx_path = _p(current_xlsx)

print(f"Current xlsx: {current_xlsx}")

if not os.path.exists(xlsx_path):
    # First-ever cup: create empty workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    wb.save(xlsx_path)
    print(f"Created empty {current_xlsx}")
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[wb.sheetnames[0]]
else:
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[wb.sheetnames[0]]

# Find rightmost Position header to determine next column
rows = list(ws.iter_rows(values_only=True))
max_pos_col = 0
if len(rows) >= 5:
    for ci, val in enumerate(rows[4]):
        if val == 'Position':
            max_pos_col = ci + 1  # 1-based

# First cup goes at col 1; subsequent cups go to the right
if max_pos_col == 0:
    col_start = 1
else:
    col_start = max_pos_col + COLS_PER_CUP
print(f"New cup at column {col_start}")

# Write cup block
ws.cell(row=2, column=col_start, value=cup_id)
ws.cell(row=3, column=col_start, value=f'Map: {cup_id} by {mapper}')
if fastest_time is not None:
    ws.cell(row=4, column=col_start + 2,
            value=f'Fastest Time: {fastest_time:.3f} by {fastest_name} in Round {fastest_round}')
ws.cell(row=5, column=col_start, value='Position')
ws.cell(row=5, column=col_start + 1, value='Name')
ws.cell(row=5, column=col_start + 2, value='Elim Time')
ws.cell(row=5, column=col_start + 3, value='Elim Round')

for i, (name, time, rnd, position) in enumerate(leaderboard):
    row = 6 + i
    ws.cell(row=row, column=col_start, value=position)
    ws.cell(row=row, column=col_start + 1, value=name)
    if time == 'DNF':
        ws.cell(row=row, column=col_start + 2, value='DNF')
    else:
        t = float(time.replace(',', '.'))
        ws.cell(row=row, column=col_start + 2, value=round(t * 1000))
    if rnd is not None:
        ws.cell(row=row, column=col_start + 3, value=rnd)

# Rename xlsx if cup_num extends the range
m = re.match(r'Eggy Cup (\d+)-(\d+)\.xlsx', current_xlsx)
if m:
    start_num, end_num = int(m.group(1)), int(m.group(2))
    if cup_num > end_num:
        new_xlsx = f'Eggy Cup {start_num}-{cup_num}.xlsx'
        new_path = _p(new_xlsx)
        wb.save(new_path)
        if new_path != xlsx_path and os.path.exists(xlsx_path):
            os.remove(xlsx_path)
        print(f"Saved as {new_xlsx} (renamed from {current_xlsx})")
        new_src = elo_src.replace(f"'{current_xlsx}'", f"'{new_xlsx}'")
        with open(elo_py_path, 'w', encoding='utf-8') as f:
            f.write(new_src)
        print(f"Updated elo_engine.py: {current_xlsx} -> {new_xlsx}")
    else:
        wb.save(xlsx_path)
        print(f"Saved to {current_xlsx}")
else:
    wb.save(xlsx_path)
    print(f"Saved to {current_xlsx}")

# ── 4. JSON backup ──
cup_json = {
    'cup': cup_id,
    'cup_num': cup_num,
    'mapper': mapper,
    'players': [
        {'pos': position, 'name': name, 'time': time, 'round': rnd}
        for name, time, rnd, position in leaderboard
    ]
}
json_path = _p(f'cup_{cup_num}.json')
with open(json_path, 'w', encoding='utf-8') as f:
    json.dump(cup_json, f, ensure_ascii=False, indent=2)
print(f"JSON backup: cup_{cup_num}.json")
print()

# ── 5. Run downstream scripts ──
def run(script, label):
    path = _p(script)
    if not os.path.exists(path):
        print(f"  SKIP: {script} not found")
        return True
    print("=" * 50)
    print(f"Running {script}...")
    print("=" * 50)
    r = subprocess.run([sys.executable, path], cwd=_dir)
    if r.returncode != 0:
        print(f"  WARNING: {label} failed!")
        return False
    return True

if not run('elo_engine.py', 'elo engine'):
    sys.exit(1)
print()
run('build_cups.py', 'cup data')
print()
run('build_fastest.py', 'fastest times')
print()

# ── 6. LTG inference if livelog exists ──
if live_log_backup and os.path.exists(live_log_backup):
    print("=" * 50)
    print("Running analyze_cup_livelog.py (LTG inference)...")
    print("=" * 50)
    r = subprocess.run([sys.executable, _p('analyze_cup_livelog.py'), str(cup_num)], cwd=_dir)
    if r.returncode != 0:
        print(f"  WARNING: analyze_cup_livelog.py failed")
else:
    print(f"(no livelog for Eggy {cup_num} — skipping LTG inference)")
print()

# ── 7. Summary ──
print("=" * 50)
print(f"{cup_id} COMPLETE")
print(f"  Players: {len(leaderboard)}")
print(f"  Winner: {winner}")
if fastest_time:
    print(f"  Fastest: {fastest_time:.3f} by {fastest_name}")
print(f"  Columns: {col_start}-{col_start+3}")
print("=" * 50)
